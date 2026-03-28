from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "nRED TABLES"

# ── Colour palette per category ──────────────────────────────────────────────
CAT_COLOURS = [
    "1F4E79",  # 1  Energy           deep blue
    "833C00",  # 2  Resistance        dark rust
    "375623",  # 3  Constraints       dark green
    "7B2D8B",  # 4  Leakage           purple
    "C00000",  # 5  Scope Rupture     red
    "404040",  # 6  GATE_SPACE        dark grey
    "1F6391",  # 7  Time              steel blue
    "0D5D2E",  # 8  Space             forest green
    "4A235A",  # 9  Cube Anatomy      deep violet
    "7D4700",  # 10 People & Roles    amber-brown
    "154360",  # 11 Documents         navy
    "117A65",  # 12 Reporting         teal
    "1A5276",  # 13 Finance           dark cyan
    "6E2F2F",  # 14 Safety & Quality  dark red-brown
    "1B4F72",  # 15 Communications    dark blue-grey
    "145A32",  # 16 Commissioning     dark emerald
]

HEADERS = [
    "Human Language Term",
    "Physics Category",
    "Sub-type",
    "Unit",
    "Gate IN Evidence",
    "Gate OUT Evidence",
    "Notes",
]

COL_WIDTHS = [32, 26, 28, 22, 30, 30, 48]

# ── All data ──────────────────────────────────────────────────────────────────
CATEGORIES = [

    ("CATEGORY 1 — ENERGY (E)\nWhat drives status change forward", [
        ("Budget", "Energy Potential", "Financial", "$", "Approved budget document", "Final account", "Total stored capacity"),
        ("Actual Cost", "Energy Released", "Financial", "$", "Invoice / claim", "Certified payment", "Energy consumed"),
        ("Labour Budget", "Energy Potential", "Human", "manhours", "Resource plan", "Timesheet summary", "Human energy capacity"),
        ("Labour Actual", "Energy Released", "Human", "manhours", "Timesheet / daily report", "Final manhour report", "Human energy consumed"),
        ("Equipment hours (plan)", "Energy Potential", "Mechanical", "machine-hours", "Equipment schedule", "—", "Mechanical energy capacity"),
        ("Equipment hours (actual)", "Energy Released", "Mechanical", "machine-hours", "Plant log", "Equipment final report", "Mechanical energy consumed"),
        ("Energy reserve / contingency", "Energy Reserve", "Financial", "$", "Risk budget allocation", "Contingency drawdown log", "Buffered potential — not yet released"),
        ("Productivity rate", "Energy Efficiency", "Ratio", "% or units/hr", "Planned rate", "Achieved rate", "Released energy / planned energy"),
        ("Momentum", "Energy Accumulation", "Trend", "units/week", "Rolling average log", "—", "Consecutive cubes of released energy"),
    ]),

    ("CATEGORY 2 — RESISTANCE (Ω)\nWhat opposes or slows status change — probabilistic, non-binary", [
        ("Weather (normal)", "Resistance", "Environmental", "probability index", "Met forecast", "Met actual log", "Routine — factored into plan"),
        ("Extreme weather", "Resistance", "Environmental", "probability index", "EOT claim", "Extension granted", "Exceeds normal allowance"),
        ("Ground conditions (unknown)", "Resistance", "Geotechnical", "probability index", "Geotech report", "Site investigation actual", "Discovered resistance, not designed"),
        ("Bureaucratic delay", "Resistance", "Organisational", "days", "Submission date", "Approval date", "Internal process friction"),
        ("Communication breakdown", "Resistance", "Informational", "frequency", "—", "NCR / RFI log", "Increases with team size"),
        ("Risk (allocated, residual)", "Resistance", "Probabilistic", "% probability", "Risk register entry", "Risk closed / triggered", "Quantified potential resistance"),
        ("Risk (unallocated)", "Resistance", "Unknown", "undefined", "—", "Incident report", "Not in model — pure resistance"),
        ("Absenteeism", "Resistance", "Human", "% labour lost", "Planned roster", "Actual attendance", "Reduces available energy"),
        ("Material shortage", "Resistance", "Supply chain", "days delay", "Delivery schedule", "Delivery actual", "Friction on energy release"),
        ("Interface dependency", "Resistance", "Systemic", "binary → probabilistic", "Interface schedule", "Interface completion", "Between cubes or between parties"),
    ]),

    ("CATEGORY 3 — CONSTRAINTS (C)\nHard boundaries — binary, non-negotiable", [
        ("Regulatory approval", "Constraint", "Legal/Compliance", "binary (Y/N)", "Application lodged", "Approval certificate", "Cannot proceed without — hard gate"),
        ("Permit to work", "Constraint", "Safety", "binary (Y/N)", "PTW issued", "PTW closed", "Binary — no partial compliance"),
        ("Hold point (ITP)", "Constraint", "Quality", "binary (Y/N)", "Inspection request", "Inspection signed off", "Inspector must witness — cannot bypass"),
        ("Witness point (ITP)", "Constraint", "Quality", "binary (Y/N)", "Notification sent", "Witness or waiver", "Can proceed if witness waives"),
        ("Contractual milestone", "Constraint", "Legal", "date", "Contract executed", "Milestone certificate", "Date-locked — triggers liquidated damages"),
        ("Design freeze", "Constraint", "Technical", "binary (Y/N)", "Change freeze notice", "—", "No scope change permitted past this point"),
        ("Jurisdictional law", "Constraint", "Legal", "binary (Y/N)", "—", "—", "External — cannot be removed by project"),
        ("Safety critical sign-off", "Constraint", "Safety", "binary (Y/N)", "Safety case submitted", "Safety acceptance", "Engineering safety gate"),
        ("Environmental clearance", "Constraint", "Legal/Environmental", "binary (Y/N)", "EIS submitted", "EPA clearance letter", "Legislative constraint"),
        ("Approved drawings (IFC)", "Constraint", "Technical", "binary (Y/N)", "IFC issue", "Construction on approved rev", "Cannot build from non-IFC drawings"),
    ]),

    ("CATEGORY 4 — LEAKAGE (L)\nEnergy consumed but not advancing status — waste within the system", [
        ("Rework", "Leakage", "Quality failure", "$ or manhours", "NCR raised", "NCR closed", "Energy spent twice on same status"),
        ("Defect rectification", "Leakage", "Quality failure", "$ or manhours", "Defect list", "Defect clearance", "Post-completion rework"),
        ("Abortive work", "Leakage", "Scope error", "$", "Variation instruction", "—", "Work done, then reversed — pure leakage"),
        ("Over-engineering", "Leakage", "Design inefficiency", "$", "—", "—", "Energy beyond what status change required"),
        ("Waiting time (within cube)", "Leakage", "GATE_SPACE micro", "hours", "—", "—", "Idle time inside an active cube"),
        ("Idle plant / equipment", "Leakage", "Resource inefficiency", "machine-hours", "Plant on-hire date", "Plant off-hire date", "Energy cost with zero output"),
        ("Meetings without decisions", "Leakage", "Informational", "manhours", "—", "—", "Energy spent, no status change"),
        ("Duplicate work", "Leakage", "Coordination failure", "manhours", "—", "—", "Same task done by multiple parties"),
        ("Data re-entry", "Leakage", "Systems failure", "manhours", "—", "—", "Digital leakage — no physical output"),
    ]),

    ("CATEGORY 5 — SCOPE RUPTURE (R)\nA status change that breaks the boundary of the cube — system boundary violation", [
        ("Scope change (approved)", "Rupture → new cube", "Formal variation", "$ / scope unit", "Variation order signed", "New cube created", "Rupture contained — absorbed into new cube"),
        ("Scope creep (unapproved)", "Rupture", "Informal variation", "$ / manhours", "—", "—", "Energy leaving cube boundary unrecorded"),
        ("Design change post-IFC", "Rupture", "Technical boundary", "binary", "Change request", "Revised IFC issued", "Rupture of design constraint boundary"),
        ("Client instruction (undirected)", "Rupture", "Contractual boundary", "binary", "Instruction received", "Variation lodged", "May or may not be compensable"),
        ("Discovery event", "Rupture", "Unknown boundary", "undefined", "—", "EOT / variation claim", "Unplanned rupture — cube boundaries violated"),
        ("Technology change", "Rupture", "Innovation boundary", "$", "—", "Business case", "Replaces existing cube structure"),
    ]),

    ("CATEGORY 6 — GATE_SPACE\nVisible idle resistance — cubes that exist but carry zero status-change energy", [
        ("Float / buffer", "GATE_SPACE", "Planned idle", "weeks", "Programme float", "—", "Intentional — planned idle resistance"),
        ("Delay (unplanned)", "GATE_SPACE", "Unplanned idle", "weeks", "Last active cube date", "Next active cube date", "Unintentional — resistance made visible"),
        ("Mobilisation gap", "GATE_SPACE", "Setup idle", "weeks", "Contract award", "First site activity", "Necessary but zero-output period"),
        ("Procurement lead time", "GATE_SPACE", "Supply idle", "weeks", "Order placed", "Goods on site", "Cube exists — nothing moves inside it"),
        ("Decision lag", "GATE_SPACE", "Governance idle", "weeks", "Submission date", "Decision date", "Resistance internal to client organisation"),
        ("Seasonal shutdown", "GATE_SPACE", "Planned idle", "weeks", "Last working day", "Return to site", "Scheduled gap — zero energy flow"),
    ]),

    ("CATEGORY 7 — TIME (t)\nThe X-axis — the only dimension that cannot be reversed", [
        ("Week", "District Cube (time unit)", "Default time slice", "7 days", "Monday 00:00", "Sunday 23:59", "Default X-unit — can be day/fortnight"),
        ("Programme start date", "t₀", "Origin", "date", "Contract / kickoff", "—", "First cube in X-axis"),
        ("Programme end date", "t_x", "Target", "date", "Contract / milestone", "Completion certificate", "Final cube boundary"),
        ("Calendar float", "GATE_SPACE", "Planned idle time", "weeks", "Programme baseline", "—", "Intentional empty cubes"),
        ("Extension of time (EOT)", "t_x shift", "Time boundary displacement", "weeks", "EOT claim", "EOT grant", "Contractual t_x pushed forward"),
        ("Delay (unplanned)", "Resistance → GATE_SPACE", "Unplanned displacement", "weeks", "Last active cube", "Next active cube", "Unintentional t_x pressure"),
        ("Deadline", "Constraint (time)", "Hard boundary", "date", "Contract", "—", "Cannot shift without contractual consequence"),
        ("Monday COMMIT", "Time boundary event", "Weekly decision gate", "binary", "Week under review", "COMMIT executed", "The moment status is locked in history"),
        ("Baseline programme", "Energy path plan", "Reference trajectory", "date range per cube", "Approved programme", "—", "The planned A→B path"),
        ("Revised programme", "Energy path update", "Re-planned trajectory", "date range per cube", "Revision submitted", "Revision approved", "New planned path after displacement"),
    ]),

    ("CATEGORY 8 — SPACE (X, Y, Z axes)\nThe three-dimensional grid in which status change occurs", [
        ("Week number", "X-axis position", "Time coordinate", "integer (week)", "—", "—", "Horizontal axis — the only irreversible one"),
        ("Department / responsibility", "Y-axis position", "Accountability coordinate", "organisational unit", "Organisation chart", "—", "Who owns the cube"),
        ("Subject / scope layer", "Z-axis position", "Scope coordinate", "work package / system", "WBS / PBS", "—", "What is being changed"),
        ("Cube", "District unit", "Atomic status unit", "1 week × 1 Y × 1 Z", "Gate IN criteria", "Gate OUT criteria", "Smallest indivisible unit of status change"),
        ("Row (Y-slice)", "Y-axis strip", "Responsibility band", "all Z at one Y, one X", "—", "—", "The unit displaced on COMMIT if not done"),
        ("Layer (X-slice)", "Weekly review plane", "Monday COMMIT target", "all Y and Z at one X", "—", "—", "The plane reviewed every Monday"),
        ("Active cube", "Cube with scope", "Energy-carrying unit", "binary", "First scope entered", "Last scope exits", "Has Gate IN, Gate OUT, or Milestone"),
        ("GATE_SPACE cube", "Empty cube", "Idle resistance unit", "binary", "First active cube", "Last active cube", "Between first and last active — idle made visible"),
        ("Space filler (modes 0–4)", "Visual GATE_SPACE marker", "Display mode", "0/1/2/3/4", "—", "—", "0=none, 1=between only, 2=+front, 3=+back, 4=all"),
        ("Work package", "Z-axis scope unit", "Scope boundary", "scope description", "Scope register", "Scope completion", "Maps to one Z-layer"),
        ("Production line", "Z-axis group", "Parallel scope band", "set of Z-layers", "—", "—", "e.g. 3 parallel train lines in QTMP"),
    ]),

    ("CATEGORY 9 — CUBE ANATOMY\nThe internal structure of each cube", [
        ("Gate IN", "Entry condition", "Status precondition", "binary (Y/N)", "Prerequisite list", "Gate IN signed", "Top-left marker — must be met to start cube"),
        ("Gate OUT", "Exit condition", "Status postcondition", "binary (Y/N)", "Completion criteria", "Gate OUT signed", "Top-right marker — must be met to close cube"),
        ("Milestone S (standard)", "Internal checkpoint", "Planned progress marker", "date", "Programme", "Achievement date", "Standard milestone — schedule reference"),
        ("Milestone C (contractual)", "Constraint checkpoint", "Contractual marker", "date", "Contract schedule", "Milestone certificate", "Hard contractual — triggers payment or LD"),
        ("DONE state", "Status confirmed", "Binary completion", "0 or 100%", "—", "Monday COMMIT", "Cube turns transparent — enters history"),
        ("NOT DONE state", "Status incomplete", "Binary non-completion", "0% (treated as)", "—", "COMMIT → row shifts +1 week", "Even 98% physical progress = 0 in status terms"),
        ("Transparent cube", "Historical status", "Past light cone", "binary", "COMMIT executed", "—", "Visible border only — status locked in past"),
        ("BoA (Basis of Approval)", "Evidence store", "Status proof", "document set", "Documents linked", "—", "Links cube to external databases / evidence"),
    ]),

    ("CATEGORY 10 — PEOPLE AND ROLES\nHuman actors translated to physics functions", [
        ("Project owner / client", "Energy source", "Funds and authority", "$ + decisions", "Contract executed", "Final payment", "Origin of energy potential"),
        ("Contractor", "Energy converter", "Executes status change", "manhours + $", "Contract award", "Practical completion", "Converts potential energy to released energy"),
        ("Subcontractor", "Secondary energy converter", "Partial scope", "manhours + $", "Subcontract award", "Sub-completion", "Nested converter"),
        ("Designer / engineer", "Constraint definer", "Sets technical boundaries", "drawings + specs", "Design brief", "IFC issue", "Defines what the cube must achieve"),
        ("Regulator", "Constraint authority", "Enforces hard boundaries", "binary decisions", "Application lodged", "Approval issued", "External — cannot be negotiated away"),
        ("Inspector / certifier", "Gate keeper", "Validates Gate OUT", "binary sign-off", "Inspection request", "Certificate issued", "Gate OUT cannot open without this role"),
        ("Project manager", "Energy coordinator", "Routes and monitors energy", "—", "—", "—", "Does NOT change status — coordinates those who do"),
        ("Sponsor", "Energy authority", "Releases energy reserves", "$ approval", "Business case", "Budget approval", "Controls contingency / Energy Reserve"),
        ("Stakeholder", "Resistance source", "Adds constraint or friction", "qualitative", "Engagement plan", "Sign-off / acceptance", "Can be managed but not eliminated"),
    ]),

    ("CATEGORY 11 — DOCUMENTS AND DELIVERABLES\nInformation objects translated to physics functions", [
        ("Contract", "Constraint set", "Legal energy boundary", "binary", "Signed contract", "—", "Defines what energy must achieve and by when"),
        ("Specification", "Constraint set", "Technical boundary", "binary", "Issued for construction", "—", "Defines what DONE looks like"),
        ("Drawing (IFC)", "Constraint (active)", "Construction boundary", "revision number", "IFC stamp", "—", "Cannot build from non-IFC"),
        ("Drawing (non-IFC)", "Leakage risk", "Uncontrolled constraint", "revision number", "—", "—", "Energy released against wrong boundary = leakage"),
        ("Programme / schedule", "Energy path plan", "Planned trajectory", "date-cube map", "Approved programme", "—", "Maps energy to cubes across X-axis"),
        ("Risk register", "Resistance catalogue", "Probabilistic resistance map", "probability × impact", "Risk identification", "Risk closed/triggered", "Quantified resistance inventory"),
        ("NCR (non-conformance report)", "Leakage record", "Quality failure evidence", "binary", "NCR raised", "NCR closed", "Documents energy that did not achieve status"),
        ("RFI (request for information)", "Resistance event", "Information gap", "days open", "RFI submitted", "RFI answered", "Information resistance — delays energy flow"),
        ("Variation order", "Rupture document", "Scope boundary change", "$ / scope", "VO submitted", "VO approved", "Formalises rupture — creates new cube"),
        ("Inspection test plan (ITP)", "Constraint map", "Quality gate schedule", "hold/witness/review", "ITP approved", "ITP complete", "Maps Gate IN/OUT requirements for each cube"),
        ("Completion certificate", "Gate OUT (final)", "Status change confirmed", "binary", "Application", "Certificate issued", "Ultimate Gate OUT for the entire programme"),
    ]),

    ("CATEGORY 12 — REPORTING AND MONITORING\nMeasurement of energy flow and resistance — observational layer", [
        ("Progress report", "Energy flow measurement", "Periodic status snapshot", "% or binary", "Reporting period start", "Report issued", "Measures energy released vs planned"),
        ("S-curve", "Energy trajectory", "Cumulative energy graph", "$ or manhours vs time", "Baseline programme", "—", "Planned vs actual energy release over X-axis"),
        ("Earned value (EV)", "Energy released (confirmed)", "Value of completed work", "$", "Work scope defined", "DONE confirmed", "Budget × % physically complete"),
        ("Planned value (PV)", "Energy scheduled", "Expected energy release", "$", "Baseline programme", "—", "What should have been released by t"),
        ("Actual cost (AC)", "Energy consumed", "Cost incurred", "$", "Invoices / timesheets", "—", "Energy spent regardless of status achieved"),
        ("Schedule performance index (SPI)", "Energy efficiency ratio", "EV / PV", "ratio (1.0 = on track)", "—", "—", "<1.0 = resistance exceeding plan"),
        ("Cost performance index (CPI)", "Energy efficiency ratio", "EV / AC", "ratio (1.0 = on track)", "—", "—", "<1.0 = leakage exceeding plan"),
        ("Forecast at completion (FAC)", "Projected total energy", "Energy path projection", "$", "Current CPI", "—", "AC + remaining budget adjusted for resistance"),
        ("Dashboard / KPI report", "Resistance indicator", "Management visibility", "various", "—", "—", "Makes resistance and leakage visible to authority"),
        ("Look-ahead programme", "Near-term energy plan", "3–6 week horizon", "cube map", "Weekly COMMIT", "—", "Short-range energy routing — feeds Monday COMMIT"),
        ("Variance report", "Resistance record", "Deviation from plan", "$ / days", "Baseline", "Actual", "Documents where resistance exceeded planned"),
        ("GATE_SPACE ratio", "Idle resistance index", "% empty cubes between active", "%", "First active cube", "Last active cube", "e.g. 47% in QTMP — not efficiency, but visible idle"),
    ]),

    ("CATEGORY 13 — FINANCE AND PROCUREMENT\nEnergy acquisition, storage, and transfer mechanisms", [
        ("Tender / bid", "Energy potential offer", "Proposed energy commitment", "$", "RFT issued", "Tender submitted", "Offer to convert energy at a price"),
        ("Contract sum", "Energy potential (contracted)", "Committed energy capacity", "$", "Contract executed", "—", "Total stored energy authorised for release"),
        ("Provisional sum", "Energy reserve (undefined)", "Unallocated energy capacity", "$", "Contract executed", "PS instruction issued", "Placeholder — scope not yet defined"),
        ("Prime cost item", "Energy reserve (defined)", "Allocated but unpriced", "$", "Contract executed", "PC instruction issued", "Known item, unknown supplier price"),
        ("Payment claim", "Energy release request", "Released energy invoice", "$", "Work done", "Claim submitted", "Contractor requesting energy transfer"),
        ("Progress payment", "Energy transfer", "Partial energy release", "$", "Claim certified", "Payment made", "Energy flowing from owner to contractor"),
        ("Retention", "Energy hold", "Withheld energy reserve", "% of claim", "Contract executed", "Defects liability end", "Held against future leakage / defect risk"),
        ("Performance bond", "Resistance buffer", "Counterparty risk cover", "$ or %", "Contract executed", "Bond released", "Financial resistance absorber"),
        ("Liquidated damages (LD)", "Resistance penalty", "Time boundary breach cost", "$/day", "Contract executed", "EOT or LD triggered", "Cost of failing to meet constraint (t_x)"),
        ("Variation claim", "Rupture energy request", "Extra energy for scope rupture", "$", "VO submitted", "VO certified", "Energy required for new cube created by rupture"),
        ("Contingency drawdown", "Energy reserve release", "Reserve → active energy", "$", "Risk event triggered", "Expenditure approved", "Reserve energy converted to released energy"),
        ("Final account", "Total energy reconciliation", "Lifecycle energy summary", "$", "Practical completion", "Final certificate", "Complete energy audit for the programme"),
    ]),

    ("CATEGORY 14 — SAFETY AND QUALITY\nConstraint enforcement and leakage prevention systems", [
        ("Safety management plan", "Constraint framework", "Safety boundary definition", "binary (approved Y/N)", "Plan submitted", "Plan approved", "Defines safety constraints for all cubes"),
        ("Safe work method statement (SWMS)", "Constraint (activity)", "Activity-level safety boundary", "binary", "SWMS prepared", "SWMS signed by workers", "Gate IN for high-risk activities"),
        ("Permit to work (PTW)", "Constraint (hard gate)", "Binary activity permission", "binary", "PTW requested", "PTW issued + closed", "Cannot start without — hard Gate IN"),
        ("Toolbox talk", "Resistance reduction", "Team information briefing", "frequency", "—", "Attendance record", "Reduces human resistance / miscommunication"),
        ("LTIFR", "Resistance index", "Lost time injury rate", "injuries per 200k hrs", "—", "—", "Measures human resistance to safe energy flow"),
        ("Incident report", "Resistance/Leakage record", "Unplanned energy loss event", "binary", "Incident occurred", "Report closed", "Documents resistance or leakage event"),
        ("Near miss report", "Resistance warning", "Near-failure signal", "binary", "Near miss occurred", "Report closed", "Early resistance signal — leading indicator"),
        ("Quality management plan", "Constraint framework", "Quality boundary definition", "binary", "Plan submitted", "Plan approved", "Defines what DONE means for all cubes"),
        ("Inspection test plan (ITP)", "Constraint map", "Gate IN/OUT schedule", "hold/witness/review", "ITP approved", "ITP complete", "Maps binary gates to specific cube activities"),
        ("Hold point", "Constraint (hard gate)", "Binary inspection gate", "binary", "Inspection request", "Inspector sign-off", "Cannot proceed — Gate OUT locked by third party"),
        ("Witness point", "Constraint (soft gate)", "Binary or waivable gate", "binary", "Notification issued", "Witness or waiver", "Can proceed if inspector waives — softer Gate OUT"),
        ("NCR (non-conformance)", "Leakage record", "Quality failure evidence", "binary", "NCR raised", "NCR closed", "Energy consumed but status not achieved"),
        ("Defect", "Leakage (post-completion)", "Post-Gate OUT failure", "binary", "Defect identified", "Defect rectified", "Gate OUT was incorrectly closed — energy lost"),
        ("Audit", "Resistance measurement", "Systematic constraint check", "score / findings", "Audit scheduled", "Audit report", "Measures resistance compliance — not status"),
        ("Concession / deviation", "Constraint relaxation", "Approved boundary shift", "binary", "Request submitted", "Approved / rejected", "Constraint moved — cube boundary adjusted"),
    ]),

    ("CATEGORY 15 — COMMUNICATIONS AND INFORMATION\nInformation energy — the flow of data that enables or blocks status change", [
        ("RFI (request for information)", "Resistance event", "Information gap", "days open", "RFI submitted", "RFI answered", "Open RFI = active resistance on cube"),
        ("Submittal", "Gate IN document", "Pre-activity evidence package", "binary", "Submittal lodged", "Submittal approved", "Gate IN cannot open until submittal approved"),
        ("Transmittal", "Information transfer", "Document delivery record", "binary", "Document issued", "Receipt confirmed", "Tracks information energy flow"),
        ("Meeting minutes", "Decision record", "Energy direction confirmation", "binary", "Meeting held", "Minutes issued + agreed", "Undistributed minutes = information leakage"),
        ("Action item", "Micro-energy task", "Discrete work unit", "binary", "Action assigned", "Action closed", "Sub-cube level — feeds cube Gate IN/OUT"),
        ("Early warning notice", "Resistance signal", "Advance resistance alert", "binary", "Notice issued", "Response received", "Contractual mechanism to flag resistance early"),
        ("Claim notice", "Rupture signal", "Boundary breach notification", "binary", "Event occurred", "Notice served", "Contractual trigger for variation / EOT"),
        ("Correspondence register", "Information flow log", "Communication audit trail", "count / days", "—", "—", "Measures information resistance over time"),
        ("BIM model", "3D constraint model", "Digital scope boundary", "revision", "BIM brief", "Model federated", "Digital definition of what cubes must achieve"),
        ("Data room / document control", "BoA evidence store", "Status proof repository", "document count", "Project start", "Final archive", "Feeds BoA — links evidence to cube completion"),
    ]),

    ("CATEGORY 16 — COMMISSIONING AND HANDOVER\nThe final status change — from constructed to operational", [
        ("Commissioning plan", "Energy path plan", "Pre-operational status map", "binary", "Plan approved", "—", "Maps final cubes before handover"),
        ("Pre-commissioning", "Gate IN (system)", "System-level readiness check", "binary", "Construction complete", "Pre-comm sign-off", "Gate IN for commissioning cube"),
        ("Commissioning", "Status change (final)", "Operational status activation", "binary", "Pre-comm complete", "Commissioning cert", "The cube where built → operational"),
        ("Factory acceptance test (FAT)", "Gate OUT (manufacture)", "Off-site status confirmation", "binary", "Test schedule", "FAT certificate", "Gate OUT for manufacturing cube"),
        ("Site acceptance test (SAT)", "Gate OUT (installation)", "On-site status confirmation", "binary", "FAT certificate", "SAT certificate", "Gate OUT for installation cube"),
        ("Practical completion", "Gate OUT (programme)", "Construction status confirmed", "binary", "PC application", "PC certificate", "Programme-level Gate OUT"),
        ("Defects liability period (DLP)", "Post-completion resistance", "Residual leakage window", "months", "PC certificate", "DLP expiry", "Period where leakage (defects) still expected"),
        ("Final completion", "Gate OUT (contract)", "Contract status confirmed", "binary", "DLP expiry", "Final certificate", "Absolute Gate OUT — contract closed"),
        ("As-built drawings", "Historical record", "Actual constraint map", "revision", "Construction complete", "As-builts approved", "Documents what status was actually achieved"),
        ("O&M manuals", "Operational energy guide", "Future energy instructions", "document set", "Handover", "Accepted by client", "Guides future status changes in operational phase"),
        ("Training (operator)", "Energy transfer (knowledge)", "Operational energy handover", "manhours", "Manuals available", "Training complete", "Human energy to operate the new status"),
        ("Spare parts schedule", "Energy reserve (operational)", "Future resistance buffer", "item list", "Handover", "Accepted", "Stored energy for future maintenance cubes"),
    ]),
]

# ── Styles ────────────────────────────────────────────────────────────────────
thin = Side(style='thin', color='CCCCCC')
med  = Side(style='medium', color='888888')

def hdr_border():
    return Border(left=med, right=med, top=med, bottom=med)

def cell_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def cat_fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def white_fill():
    return PatternFill("solid", fgColor="FFFFFF")

def alt_fill():
    return PatternFill("solid", fgColor="F5F5F5")

# ── Column widths ─────────────────────────────────────────────────────────────
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Title row ─────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 36
title_cell = ws.cell(row=1, column=1,
    value="nRED TABLES — Physics Translation Dictionary\nnREDCube™ PRINCIPIA · Universal Principles of Status Change Management")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
title_cell.font  = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
title_cell.fill  = cat_fill("1A1A2E")
title_cell.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)

# ── Formula row ───────────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 22
formula_cell = ws.cell(row=2, column=1,
    value="Core Formula:  S(A→B) = Δ(E·t) / ΔΩ        "
          "where  E = Energy/Effort   ·   t = time   ·   Ω = Impedance/Resistance")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
formula_cell.font  = Font(name='Calibri', italic=True, size=11, color='C00000')
formula_cell.fill  = cat_fill("F2F2F2")
formula_cell.alignment = Alignment(horizontal='center', vertical='center')

current_row = 4

for cat_idx, (cat_title, rows) in enumerate(CATEGORIES):
    hex_col = CAT_COLOURS[cat_idx]

    # Category header
    ws.row_dimensions[current_row].height = 32
    cat_cell = ws.cell(row=current_row, column=1, value=cat_title)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=7)
    cat_cell.font  = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    cat_cell.fill  = cat_fill(hex_col)
    cat_cell.alignment = Alignment(horizontal='left', vertical='center',
                                   indent=1, wrap_text=True)
    current_row += 1

    # Column headers
    ws.row_dimensions[current_row].height = 28
    for col_idx, hdr in enumerate(HEADERS, 1):
        c = ws.cell(row=current_row, column=col_idx, value=hdr)
        c.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill      = cat_fill(hex_col)
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        c.border    = hdr_border()
    current_row += 1

    # Data rows
    for row_idx, row_data in enumerate(rows):
        ws.row_dimensions[current_row].height = 42
        fill = white_fill() if row_idx % 2 == 0 else alt_fill()
        for col_idx, value in enumerate(row_data, 1):
            c = ws.cell(row=current_row, column=col_idx, value=value)
            c.font      = Font(name='Calibri', size=10)
            c.fill      = fill
            c.alignment = Alignment(horizontal='left', vertical='center',
                                    wrap_text=True, indent=1)
            c.border    = cell_border()
            # Bold the term column
            if col_idx == 1:
                c.font = Font(name='Calibri', size=10, bold=True)
        current_row += 1

    current_row += 1  # blank row between categories

# ── Freeze top rows ───────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Auto-filter on header of first category (optional) ───────────────────────
# Not set globally as categories have individual headers

output_path = r"c:\Users\Daddy\nREDCube.org\nRED_TABLES_v0.2.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
